from django.shortcuts import render
import openpyxl


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        # for col in worksheet.iter_cols():
        #     col_data = list()
        #     for cell in col:
        #         col_data.append(str(cell.value))
        #     excel_data.append(col_data)

        # column name mapping
        # colnames = ['first_name', 'last_name', 'national_id', 'email', 'phone_no']

        # col_indices = {n for n, cell in enumerate(worksheet.rows[0]) if cell.value in colnames}
        # print(col_indices.value)

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(row_data)
                print(row[0].value)
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data":excel_data})
